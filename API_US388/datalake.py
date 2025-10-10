import requests
import io
from io import BytesIO
import uuid
import boto3
from openpyxl import load_workbook
from botocore.exceptions import ClientError
import os
from pathlib import Path
import pandas as pd


def download_file_from_s3(bucket_name, file_name):
    try:
        session = boto3.Session(region_name='eu-west-1')
        s3 = session.client('s3',aws_access_key_id="AKIAUJXLE2HEM4YQVOSA",aws_secret_access_key="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")

        response = s3.get_object(Bucket=bucket_name, Key=file_name)
        excel_content = response['Body'].read()

        excel_data = BytesIO(excel_content)
        workbook = load_workbook(excel_data)
        sheet = workbook.active

        lines = []
        for row in sheet.iter_rows(values_only=True):
            lines.append(row)

        print(f"Excel file '{file_name}' downloaded and parsed successfully from bucket '{bucket_name}'.")
        print(str(lines[1]))
        return lines
    except Exception as e:
        print(f"Error parsing Excel file '{file_name}' from bucket '{bucket_name}': {e}")
        return None

    
def upload_file_to_s3(file_path,s3_bucket,s3_key):
    try:
        session = boto3.Session(region_name='eu-west-1')
        s3 = session.client('s3', aws_access_key_id="AKIAUJXLE2HEM4YQVOSA",aws_secret_access_key="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")
        print("filepath: " ,str(file_path))
        s3.upload_file(file_path, s3_bucket, s3_key)
        url = generate_download_url(s3_bucket, s3_key)
    except Exception as e:
        print(f"An error occurred: {e}")
    return url

def generate_download_url(bucket_name, object_name, expiration=3600):
    session = boto3.Session(region_name='eu-west-1')
    s3_client = session.client('s3',aws_access_key_id="AKIAUJXLE2HEM4YQVOSA",aws_secret_access_key="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")
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': object_name},
            ExpiresIn=expiration
        )
    except ClientError as e:
        print(f"Error generating presigned URL: {e}")
        return None

    return url

def put_image_in_s3(media,bucket_name,object_key):
    s3 = boto3.client('s3', aws_access_key_id="AKIAUJXLE2HEM4YQVOSA",aws_secret_access_key="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",region_name='eu-west-1')

    try:
        s3.put_object(Bucket=bucket_name, Key=object_key, Body=media)
        print("Media uploaded successfully to S3")
    except Exception as e:
        print(f"An error occurred: {e}")


def read_metadata_from_s3(bucket_name, file_key):
    s3 = boto3.client('s3', aws_access_key_id="AKIAUJXLE2HEM4YQVOSA",aws_secret_access_key="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",region_name='eu-west-1')
    obj = s3.get_object(Bucket=bucket_name, Key=file_key)
    return pd.read_excel(BytesIO(obj['Body'].read()))